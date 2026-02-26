import io
import logging
from typing import Dict, Any, List, Optional
from openpyxl import load_workbook

logger = logging.getLogger(__name__)

class ExcelBpImportService:
    """Service for parsing Business Plan Excel files to extract input values."""

    @staticmethod
    def parse_upload(file_content: bytes, current_bp_dict: Dict[str, Any]) -> Dict[str, Any]:
        """
        Parse an uploaded Excel file and build an update dictionary for the Business Plan.
        
        Args:
            file_content: Raw bytes of the uploaded Excel file
            current_bp_dict: Dictionary representation of the current Business Plan
            
        Returns:
            Dictionary with the updated input values
        """
        wb = load_workbook(io.BytesIO(file_content), data_only=True)
        updates = {}
        
        # 1. Parse PARAMETRI sheet
        if "PARAMETRI" in wb.sheetnames:
            ws = wb["PARAMETRI"]
            params = ExcelBpImportService._parse_parametri(ws)
            if params:
                for key, val in params.items():
                    if val is not None:
                        updates[key] = val
        else:
            logger.warning("Sheet 'PARAMETRI' not found in uploaded Excel.")

        # 2. Parse TEAM sheet
        if "TEAM" in wb.sheetnames:
            ws = wb["TEAM"]
            team_composition, volume_adjustments = ExcelBpImportService._parse_team(
                ws, 
                current_bp_dict.get("team_composition", []),
                current_bp_dict.get("tows", []),
                current_bp_dict.get("volume_adjustments", {})
            )
            if team_composition:
                updates["team_composition"] = team_composition
            if volume_adjustments:
                updates["volume_adjustments"] = volume_adjustments
        else:
            logger.warning("Sheet 'TEAM' not found in uploaded Excel.")
            
        return updates

    @staticmethod
    def _parse_parametri(ws) -> Dict[str, Any]:
        """Parse the PARAMETRI sheet based on labels in Column A."""
        params = {}
        
        # Field mapping: Label prefix -> (dictionary_key, type_converter)
        field_map = {
            "Durata Contratto": ("duration_months", int),
            "Giorni/Anno per FTE": ("days_per_fte", float),
            "Tariffa Default": ("default_daily_rate", float),
            "Governance %": ("governance_pct", float),
            "Risk Contingency %": ("risk_contingency_pct", float),
            "Reuse Factor %": ("reuse_factor", float),
            "Inflazione YoY": ("inflation_pct", float),
            "Sconto Offerta %": ("discount_pct", float),
            "Margine Target %": ("target_margin_pct", float)
        }
        
        for row in range(1, ws.max_row + 1):
            label_cell = ws.cell(row=row, column=1).value
            val_cell = ws.cell(row=row, column=2).value
            
            if label_cell and isinstance(label_cell, str):
                label_cell = label_cell.strip()
                for prefix, (key, type_conv) in field_map.items():
                    if label_cell.startswith(prefix) and val_cell is not None:
                        try:
                            # Parse percentages properly
                            if prefix in ["Governance %", "Risk Contingency %", "Reuse Factor %", "Sconto Offerta %", "Margine Target %"]:
                                # Excel decimal input 0.05 or raw number 5. Backend expects decimals (0.0 - 1.0).
                                val = float(val_cell)
                                if val > 1.0:
                                    val = val / 100.0  
                                params[key] = val
                            else:
                                params[key] = type_conv(val_cell)
                        except (ValueError, TypeError) as e:
                            logger.warning(f"Failed to parse '{label_cell}' with value {val_cell}: {e}")
                            
        return params

    @staticmethod
    def _parse_team(ws, current_team: List[Dict[str, Any]], tows: List[Dict[str, Any]], current_volume_adj: Dict[str, Any]) -> tuple:
        """Parse TEAM sheet for FTE base, seniority, TOW allocations, and volume adjustments."""
        if ws.max_row < 4:
            return None, None
            
        # Find headers on row 4
        headers = [ws.cell(row=4, column=c).value for c in range(1, ws.max_column + 1)]
        if not headers or "Profilo" not in headers:
            return None, None
            
        col_idx = {h: i + 1 for i, h in enumerate(headers) if h}
        
        tow_ids = [t.get("tow_id", t.get("id")) for t in tows]
        
        new_team = []
        # Dictionary to build volume constraints
        by_profile = {}
        
        for row in range(5, ws.max_row + 1):
            profilo = ws.cell(row=row, column=col_idx.get("Profilo", 1)).value
            if not profilo:
                continue
                
            if str(profilo).strip().upper() == "TOTALE":
                break # Reached the end of data
                
            # Find the original team member by label to retain its profile_id
            orig_member = next((m for m in current_team if m.get("label") == profilo or m.get("profile_id") == profilo), None)
            profile_id = orig_member.get("profile_id", profilo) if orig_member else profilo
            
            seniority = ws.cell(row=row, column=col_idx.get("Seniority", 2)).value
            fte = ws.cell(row=row, column=col_idx.get("FTE Base", 3)).value
            
            # Read TOW allocations
            tow_allocation = {}
            for tow_id in tow_ids:
                tow_col_name = f"{tow_id} %"
                if tow_col_name in col_idx:
                    val = ws.cell(row=row, column=col_idx[tow_col_name]).value
                    try:
                        # Convert 0.5 back to 50 for the json representation
                        alloc_pct = float(val) if val is not None else 0.0
                        if alloc_pct <= 1.0 and alloc_pct > 0.0:
                            alloc_pct = alloc_pct * 100.0
                        tow_allocation[tow_id] = alloc_pct
                    except (ValueError, TypeError):
                        tow_allocation[tow_id] = 0.0
                        
            # Read volume adjustments (Fattore Rid.)
            if "Fattore Rid." in col_idx:
                factor = ws.cell(row=row, column=col_idx["Fattore Rid."]).value
                try:
                    factor_val = float(factor) if factor is not None else 1.0
                    by_profile[profile_id] = factor_val
                except (ValueError, TypeError):
                    by_profile[profile_id] = 1.0
                    
            
            new_member = {
                "profile_id": profile_id,
                "label": orig_member.get("label", profilo) if orig_member else profilo,
                "seniority": seniority if seniority else "mid",
                "fte": float(fte) if fte is not None else 0.0,
                "tow_allocation": tow_allocation
            }
            new_team.append(new_member)
            
        # Create volume_adjustments block
        new_vol_adj = dict(current_volume_adj) if current_volume_adj else {}
        periods = new_vol_adj.get("periods", [])
        if not periods:
            periods = [{"month_start": 1, "month_end": 36, "by_profile": {}, "by_tow": {}}]
            
        if periods:
            # We assume the imported factors apply to the first period for simplicity since we edited them uniformly on the TEAM sheet
            periods[0]["by_profile"] = by_profile
            new_vol_adj["periods"] = periods
            
        return new_team, new_vol_adj
